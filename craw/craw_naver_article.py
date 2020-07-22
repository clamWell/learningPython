import requests
from bs4 import BeautifulSoup


#### 파이썬을 이용해 키워드별 네이버 뉴스 목록 긁어와 언론사별로 축적해보기
news_articles = {};

keywords = input("뉴스 키워드를 입력해주세요!\n")
print("뉴스를 검색중입니다")
for page in range(10):	
	raw = requests.get('https://search.naver.com/search.naver?&where=news&query='+str(keywords)+'&start=' + str(page*10+1), headers={'User-Agent': 'Mozilla/5.0'}).text
	html = BeautifulSoup(raw, 'html.parser')
	news_list = html.select('.news .type01 > li')

	for news in news_list:
		title = news.select_one('._sp_each_title').text
		newspaper = news.select_one('._sp_each_source').text.replace('언론사 선정','')
		
		if newspaper in news_articles.keys():
			news_articles[newspaper].append(title)
		else:
			news_articles[newspaper] = [title]

for items in news_articles.items():
	print(items)


#### 키워드별 네이버 뉴스 목록을 엑셀로 저장해보기 

import openpyxl
import datetime
dt = datetime.datetime.now()
file_name = 'naver_research_'+str(keywords)+'_'+dt.strftime('%Y_%m_%d')

excel = openpyxl.Workbook()
work_sheet = excel.active 
work_sheet.title = "키워드 별 네이버 뉴스 목록"
row = 1

#sheet를 담은 변수에서 활용가능한 .cell()함수는 row,와 column의 속성값을 인자로 받고
#여기서.value 함수는 해당 위치의 셀에 값을 지정해줍니다.

for items in news_articles.items():
	work_sheet.cell(row=row, column=1).value = items[0]

	for news_title in items[1]:
		work_sheet.cell(row=row, column=2).value = news_title
		row +=1

	row+=1

excel.save(file_name+'.xlsx')
excel.close()


